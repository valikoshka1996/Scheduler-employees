import calendar
import random
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

# Список працівників
WEEKEND_WORKERS = ["Данило", "Влад", "Маша"]
WEEKDAY_WORKERS = ["Вова", "Руслан", "Валік", "Діма", "Олексій", "Юра"]

# Унікальні кольори для працівників
COLORS = {
    "Данило": "FF9999", "Влад": "FFCC99", "Маша": "FFFF99",
    "Вова": "99FF99", "Руслан": "99FFFF", "Валік": "9999FF",
    "Діма": "CC99FF", "Олексій": "FF99CC", "Юра": "CCCCCC"
}

WEEKDAYS = ["Понеділок", "Вівторок", "Середа", "Четвер", "П'ятниця", "Субота", "Неділя"]

def validate_month_year(date_str):
    """Перевіряє формат MM.YYYY."""
    return bool(re.match(r"^(0[1-9]|1[0-2])\.\d{4}$", date_str))

def get_first_day_worker(workers, prompt):
    """Пропонує користувачу вибрати першого працівника."""
    print(prompt)
    for i, worker in enumerate(workers, 1):
        print(f"{i}. {worker}")
    while True:
        choice = input("Введіть номер: ")
        if choice.isdigit() and 1 <= int(choice) <= len(workers):
            return workers[int(choice) - 1]
        print("Некоректний вибір. Спробуйте ще раз.")

def generate_schedule(month, year, additional_weekends, first_worker, first_weekday_worker):
    """Генерує графік змін за правильною логікою."""
    schedule = {}
    num_days = calendar.monthrange(year, month)[1]
    workers_by_day_type = {"будній": WEEKDAY_WORKERS.copy(), "вихідний": WEEKEND_WORKERS.copy()}
    
    weekday_queue = workers_by_day_type["будній"].copy()
    weekend_queue = workers_by_day_type["вихідний"].copy()
    
    if first_worker in weekend_queue:
        weekend_queue.remove(first_worker)
        weekend_queue.insert(2, first_worker)
    
    if first_weekday_worker:
        weekday_queue.remove(first_weekday_worker)
        weekday_queue.insert(2, first_weekday_worker)
    
    for day in range(1, num_days + 1):
        current_date = datetime(year, month, day)
        is_weekend = current_date.weekday() >= 5 or day in additional_weekends
        day_type = "вихідний" if is_weekend else "будній"
        workers_queue = weekend_queue if is_weekend else weekday_queue
        
        if len(workers_queue) < 3:
            workers_queue = workers_by_day_type[day_type].copy()
        
        schedule[day] = {
            "Третя зміна": workers_queue[0],
            "Друга зміна": workers_queue[1],
            "Перша зміна": workers_queue[2]
        }
        workers_queue.append(workers_queue.pop(0))
    
    return schedule

def save_to_excel(schedule, month, year):
    """Зберігає графік у файл Excel у горизонтальному форматі, кожна зміна займає 4 рядки."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    header = ["Дата"] + [f"{day:02}.{month:02}.{year}" for day in schedule.keys()]
    ws.append(header)
    
    ws.append(["День тижня"] + [WEEKDAYS[datetime(year, month, day).weekday()] for day in schedule.keys()])
    
    for _ in range(4):
        ws.append(["Перша зміна"] + [schedule[day]["Перша зміна"] for day in schedule.keys()])
    for _ in range(4):
        ws.append(["Друга зміна"] + [schedule[day]["Друга зміна"] for day in schedule.keys()])
    for _ in range(4):
        ws.append(["Третя зміна"] + [schedule[day]["Третя зміна"] for day in schedule.keys()])
    
    for row in ws.iter_rows(min_row=3, min_col=2):
        for cell in row:
            if cell.value in COLORS:
                cell.fill = PatternFill(start_color=COLORS[cell.value], fill_type="solid")
    
    filename = f"Графік_{month:02}_{year}.xlsx"
    wb.save(filename)
    print(f"Графік збережено у {filename}")

def main():
    """Основна функція."""
    while True:
        month_year = input("Введіть місяць та рік у форматі MM.YYYY: ")
        if validate_month_year(month_year):
            break
        print("Некоректний формат! Спробуйте ще раз.")
    
    month, year = map(int, month_year.split("."))
    additional_weekends = list(map(int, input("Введіть додаткові вихідні (через пробіл): ").split()))
    first_day = datetime(year, month, 1)
    is_weekend = first_day.weekday() >= 5 or 1 in additional_weekends
    
    first_worker = get_first_day_worker(WEEKEND_WORKERS if is_weekend else WEEKDAY_WORKERS, 
                                        "Оберіть першого працівника для третьої зміни першого дня:")
    
    first_weekday_worker = None
    if is_weekend:
        first_weekday_worker = get_first_day_worker(WEEKDAY_WORKERS, 
                                                    "Оберіть першого працівника для першої зміни першого буднього дня:")
    
    schedule = generate_schedule(month, year, additional_weekends, first_worker, first_weekday_worker)
    save_to_excel(schedule, month, year)
    
if __name__ == "__main__":
    main()
