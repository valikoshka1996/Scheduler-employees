import calendar
import random
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import messagebox, ttk

# -------------------- Основна логіка --------------------
DEFAULT_WEEKDAY = ["Вова", "Руслан", "Валік", "Діма", "Олексій", "Юра"]
DEFAULT_WEEKEND = ["Данило", "Влад", "Маша"]
WEEKDAYS = ["Понеділок", "Вівторок", "Середа", "Четвер", "П'ятниця", "Субота", "Неділя"]

def generate_color():
    return "{:06x}".format(random.randint(0, 0xFFFFFF)).upper()

def move_nth_to_first(lst, n):
    """Переміщує елемент за індексом n на перше місце у списку.
    
    Умови:
    - Якщо список містить 3 елементи – останній стає першим, зміщуючи інші.
    - Якщо список містить більше ніж 3 елементи – елемент за індексом n стає першим, а всі інші зміщуються далі.
    """
    if len(lst) <= 3:
        return [lst[-1]] + lst[:-1]
    elif len(lst) > 3:
        if not (0 <= n < len(lst)):
            raise ValueError("Неправильний індекс")
        return [lst[n]] + lst[:n] + lst[n+1:]
    else:
        return lst

def generate_schedule(month, year, additional_weekends, weekend_workers, weekday_workers, vacation_days=None):
    """
    Генерує графік змін з урахуванням відпусток.
    
    Для кожного дня обирається відповідна базова черга (будня або вихідна), яка оновлюється за допомогою move_nth_to_first.
    Потім формується "ефективна" черга – це базова черга без тих співробітників, які у відпустці (тобто, якщо поточний день
    міститься у списку їхніх відпусток). Призначення змін відбувається за порядком:
      • Перша зміна – перший елемент ефективної черги,
      • Друга зміна – другий,
      • Третя зміна – третій.
    Якщо ефективна черга містить менше 3 співробітників, використовується циклічне призначення:
      • Якщо 2 співробітники – перша: ефектив[0], друга: ефектив[1], третя: ефектив[0].
      • Якщо 1 – усі зміни отримують одного.
      • Якщо ефективна черга порожня – у всі зміни записується "Відпустка".
    
    При цьому базова черга залишається незмінною для наступних днів (тобто, якщо співробітник повертається з відпустки, він повертається на своє місце).
    """
    if vacation_days is None:
        vacation_days = {}
    schedule = {}
    num_days = calendar.monthrange(year, month)[1]

    # Початкові значення для обертання
    index_week = 3
    index_weeknd = 3
    workers_by_day_type = {
        "будній": weekday_workers.copy(),
        "вихідний": weekend_workers.copy()
    }
    weekday_queue = workers_by_day_type["будній"].copy()
    weekend_queue = workers_by_day_type["вихідний"].copy()
    
    for day in range(1, num_days + 1):
        current_date = datetime(year, month, day)
        is_weekend = current_date.weekday() >= 5 or day in additional_weekends
        
        # Вибір базової черги для даного дня
        current_queue = weekend_queue if is_weekend else weekday_queue
        
        # Формування ефективної черги: видаляємо тих, хто у відпустці (для поточного дня)
        effective_queue = [worker for worker in current_queue if day not in vacation_days.get(worker, [])]
        
        # Призначення змін за порядком із використанням ефективної черги
        if len(effective_queue) >= 3:
            shifts = {
                "Перша зміна": effective_queue[0],
                "Друга зміна": effective_queue[1],
                "Третя зміна": effective_queue[2]
            }
        elif len(effective_queue) == 2:
            shifts = {
                "Перша зміна": effective_queue[0],
                "Друга зміна": effective_queue[1],
                "Третя зміна": effective_queue[0]
            }
        elif len(effective_queue) == 1:
            shifts = {
                "Перша зміна": effective_queue[0],
                "Друга зміна": effective_queue[0],
                "Третя зміна": effective_queue[0]
            }
        else:
            shifts = {"Перша зміна": "Відпустка", "Друга зміна": "Відпустка", "Третя зміна": "Відпустка"}
        schedule[day] = shifts
        
        # Оновлення базової черги за оригінальною логікою (без врахування відпусток)
        if is_weekend:
            if len(weekend_workers) == 2:
                index_weeknd = 1
            elif len(weekend_workers) == 1:
                index_weeknd = 0
            index = index_weeknd
            weekend_queue = move_nth_to_first(current_queue, index)
            if index_weeknd == len(weekend_workers) - 1:
                index_weeknd = len(weekend_workers) - 1
            elif len(weekend_workers) == 1:
                index_weeknd = 0
            else:
                index_weeknd += 1
        else:
            if len(weekday_workers) == 2:
                index_week = 1
            elif len(weekday_workers) == 1:
                index_week = 0
            index = index_week
            weekday_queue = move_nth_to_first(current_queue, index)
            if index_week == len(weekday_workers) - 1:
                index_week = len(weekday_workers) - 1
            elif len(weekday_workers) == 1:
                index_week = 0
            else:
                index_week += 1
    
    return schedule

def save_to_excel(schedule, month, year):
    wb = openpyxl.Workbook()
    ws = wb.active
    header = ["Дата"] + [f"{day:02}.{month:02}.{year}" for day in sorted(schedule.keys())]
    ws.append(header)
    ws.append(["День тижня"] + [WEEKDAYS[datetime(year, month, day).weekday()] for day in sorted(schedule.keys())])
    
    all_workers = {worker for day in schedule.values() for worker in day.values() if worker not in ["Немає", "Відпустка"]}
    COLORS = {worker: generate_color() for worker in all_workers}
    
    shift_names = ["Перша зміна", "Друга зміна", "Третя зміна"]
    for shift in shift_names:
        for _ in range(4):
            ws.append([shift] + [schedule[day][shift] for day in sorted(schedule.keys())])
    
    for row in ws.iter_rows(min_row=3, min_col=2):
        for cell in row:
            if cell.value in COLORS:
                cell.fill = PatternFill(start_color=COLORS[cell.value], fill_type="solid")
    
    filename = f"Графік_{month:02}_{year}.xlsx"
    wb.save(filename)
    return filename

# -------------------- Інтерфейс --------------------
def update_weekday_start_dropdown():
    lines = text_weekday.get("1.0", tk.END).strip().splitlines()
    combobox_weekday_start['values'] = lines
    combobox_weekday_start.set(lines[0] if lines else "")

def update_weekend_start_dropdown():
    lines = text_weekend.get("1.0", tk.END).strip().splitlines()
    combobox_weekend_start['values'] = lines
    combobox_weekend_start.set(lines[0] if lines else "")

# Глобальний словник для зберігання відпусток: ключ – ім'я, значення – список днів
global_vacation_days = {}

def update_vacation_display():
    text_vac_display.config(state="normal")
    text_vac_display.delete("1.0", tk.END)
    for emp, days in global_vacation_days.items():
        text_vac_display.insert(tk.END, f"{emp}: {', '.join(map(str, sorted(days)))}\n")
    text_vac_display.config(state="disabled")

def open_vacation_window():
    # Вікно для додавання відпусток (збільшене, щоб кнопки було видно)
    vac_win = tk.Toplevel(root)
    vac_win.title("Додати відпустку")
    vac_win.geometry("400x300")
    
    # Тимчасовий словник для роботи у цьому вікні
    temp_vacations = global_vacation_days.copy()
    
    def update_employee_list():
        weekday_emps = text_weekday.get("1.0", tk.END).strip().splitlines()
        weekend_emps = text_weekend.get("1.0", tk.END).strip().splitlines()
        employees = list(dict.fromkeys(weekday_emps + weekend_emps))
        combo_employee['values'] = employees
        if employees:
            combo_employee.set(employees[0])
    
    # Розташування елементів за допомогою grid (розміщено поруч)
    ttk.Label(vac_win, text="Співробітник:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
    combo_employee = ttk.Combobox(vac_win, state="readonly")
    combo_employee.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    update_employee_list()
    
    ttk.Label(vac_win, text="Дні відпустки (через кому):").grid(row=1, column=0, padx=5, pady=5, sticky="w")
    entry_vac_days = ttk.Entry(vac_win)
    entry_vac_days.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
    
    listbox_vac = tk.Listbox(vac_win, width=40)
    listbox_vac.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")
    
    def update_vacation_listbox():
        listbox_vac.delete(0, tk.END)
        for emp, days in temp_vacations.items():
            listbox_vac.insert(tk.END, f"{emp}: {', '.join(map(str, sorted(days)))}")
    
    def add_vacation_to_temp():
        emp = combo_employee.get().strip()
        days_str = entry_vac_days.get().strip()
        if not emp or not days_str:
            messagebox.showerror("Помилка", "Введіть ім'я співробітника та дні відпустки")
            return
        days = []
        for part in days_str.split(","):
            part = part.strip()
            if part.isdigit():
                days.append(int(part))
        if not days:
            messagebox.showerror("Помилка", "Введіть коректні дні відпустки")
            return
        if emp in temp_vacations:
            temp_vacations[emp].extend(days)
            temp_vacations[emp] = sorted(list(set(temp_vacations[emp])))
        else:
            temp_vacations[emp] = sorted(days)
        update_vacation_listbox()
        entry_vac_days.delete(0, tk.END)
    
    def save_vacation_data():
        global global_vacation_days
        global_vacation_days = temp_vacations.copy()
        update_vacation_display()
        messagebox.showinfo("Успіх", "Відпустки збережено!")
        vac_win.destroy()
    
    btn_add = ttk.Button(vac_win, text="Додати в список", command=add_vacation_to_temp)
    btn_add.grid(row=3, column=0, padx=5, pady=5, sticky="ew")
    btn_save = ttk.Button(vac_win, text="Зберегти відпустку", command=save_vacation_data)
    btn_save.grid(row=3, column=1, padx=5, pady=5, sticky="ew")
    btn_close = ttk.Button(vac_win, text="Закрити", command=vac_win.destroy)
    btn_close.grid(row=4, column=0, columnspan=2, padx=5, pady=5, sticky="ew")
    
    vac_win.columnconfigure(0, weight=1)
    vac_win.columnconfigure(1, weight=1)
    vac_win.rowconfigure(2, weight=1)

def generate_schedule_from_gui():
    try:
        year = int(combobox_year.get().strip())
        month = int(combobox_month.get().strip())
    except ValueError:
        messagebox.showerror("Помилка", "Виберіть коректний рік та місяць")
        return
    additional_weekends = []
    if enable_additional_var.get():
        for day, var in additional_days_vars.items():
            if var.get() == 1:
                additional_weekends.append(day)
    weekend_workers = text_weekend.get("1.0", tk.END).strip().splitlines()
    weekday_workers = text_weekday.get("1.0", tk.END).strip().splitlines()
    weekday_start = combobox_weekday_start.get().strip()
    weekend_start = combobox_weekend_start.get().strip()
    if weekday_start in weekday_workers:
        idx = weekday_workers.index(weekday_start)
        weekday_workers = weekday_workers[idx:] + weekday_workers[:idx]
    if weekend_start in weekend_workers:
        idx = weekend_workers.index(weekend_start)
        weekend_workers = weekend_workers[idx:] + weekend_workers[:idx]
    schedule = generate_schedule(month, year, additional_weekends, weekend_workers, weekday_workers, global_vacation_days)
    try:
        filename = save_to_excel(schedule, month, year)
        messagebox.showinfo("Успіх", f"Графік збережено у {filename}")
    except Exception as e:
        messagebox.showerror("Помилка", f"Не вдалося зберегти графік: {e}")

def update_additional_days_grid(*args):
    for widget in frame_additional_days.winfo_children():
        widget.destroy()
    additional_days_vars.clear()
    try:
        year = int(combobox_year.get().strip())
        month = int(combobox_month.get().strip())
    except ValueError:
        return
    cal = calendar.monthcalendar(year, month)
    days_of_week = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Нд"]
    for col, day_name in enumerate(days_of_week):
        lbl = ttk.Label(frame_additional_days, text=day_name, width=3, anchor="center")
        lbl.grid(row=0, column=col, padx=1, pady=1)
    row_offset = 1
    for r, week in enumerate(cal):
        for c, day in enumerate(week):
            if day == 0:
                lbl = ttk.Label(frame_additional_days, text="", width=3)
                lbl.grid(row=r+row_offset, column=c, padx=1, pady=1)
            else:
                if c < 5:
                    var = tk.IntVar(value=0)
                    chk = ttk.Checkbutton(frame_additional_days, text=str(day), variable=var)
                    chk.grid(row=r+row_offset, column=c, padx=1, pady=1)
                    additional_days_vars[day] = var
                else:
                    lbl = ttk.Label(frame_additional_days, text=str(day), width=3, foreground="gray")
                    lbl.grid(row=r+row_offset, column=c, padx=1, pady=1)

def on_enable_additional_toggle():
    if enable_additional_var.get():
        frame_additional_days.grid()
        update_additional_days_grid()
    else:
        frame_additional_days.grid_remove()

# -------------------- Головне вікно --------------------
if __name__ == '__main__':
    root = tk.Tk()
    root.title("Генератор графіку змін")
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Custom.TFrame", background="#f0f0f0", relief="groove", borderwidth=2, padding=10)
    additional_days_vars = {}
    
    frame_top = ttk.Frame(root)
    frame_top.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
    ttk.Label(frame_top, text="Рік:").grid(row=0, column=0, padx=2, pady=2, sticky="w")
    combobox_year = ttk.Combobox(frame_top, values=[str(y) for y in range(2025, 2031)], state="readonly", width=5)
    combobox_year.grid(row=0, column=1, padx=2, pady=2)
    combobox_year.current(0)
    ttk.Label(frame_top, text="Місяць:").grid(row=0, column=2, padx=2, pady=2, sticky="w")
    combobox_month = ttk.Combobox(frame_top, values=[str(m) for m in range(1, 13)], state="readonly", width=3)
    combobox_month.grid(row=0, column=3, padx=2, pady=2)
    combobox_month.current(0)
    enable_additional_var = tk.IntVar(value=0)
    chk_enable_additional = ttk.Checkbutton(frame_top, text="Додаткові вихідні?", variable=enable_additional_var,
                                            command=on_enable_additional_toggle)
    chk_enable_additional.grid(row=0, column=4, padx=2, pady=2)
    btn_generate = ttk.Button(frame_top, text="Генерувати графік", command=generate_schedule_from_gui)
    btn_generate.grid(row=0, column=5, padx=10, pady=2, sticky="e")
    
    # Кнопка для відкриття вікна додавання відпусток
    btn_open_vac = ttk.Button(frame_top, text="Додати відпустку", command=open_vacation_window)
    btn_open_vac.grid(row=0, column=6, padx=10, pady=2)
    
    frame_additional_days = ttk.Frame(root)
    frame_additional_days.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
    frame_additional_days.grid_remove()
    combobox_year.bind("<<ComboboxSelected>>", update_additional_days_grid)
    combobox_month.bind("<<ComboboxSelected>>", update_additional_days_grid)
    
    frame_workers = ttk.Frame(root)
    frame_workers.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
    
    frm_weekend = ttk.Frame(frame_workers, style="Custom.TFrame")
    frm_weekend.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")
    ttk.Label(frm_weekend, text="Співробітники (вихідні):").grid(row=0, column=0, padx=5, pady=(5,2), sticky="w")
    default_weekend_var = tk.IntVar(value=0)
    chk_default_weekend = ttk.Checkbutton(frm_weekend, text="Default вихідні", variable=default_weekend_var,
                                           command=lambda: (text_weekend.delete("1.0", tk.END),
                                                            text_weekend.insert(tk.END, "\n".join(DEFAULT_WEEKEND)),
                                                            update_weekend_start_dropdown()))
    chk_default_weekend.grid(row=1, column=0, padx=5, pady=2, sticky="w")
    text_weekend = tk.Text(frm_weekend, width=30, height=8)
    text_weekend.grid(row=2, column=0, padx=5, pady=5)
    ttk.Label(frm_weekend, text="Стартовий:").grid(row=3, column=0, padx=5, pady=(5,2), sticky="w")
    combobox_weekend_start = ttk.Combobox(frm_weekend, state="readonly", width=27)
    combobox_weekend_start.grid(row=4, column=0, padx=5, pady=2, sticky="w")
    update_weekend_start_dropdown()
    
    frm_weekday = ttk.Frame(frame_workers, style="Custom.TFrame")
    frm_weekday.grid(row=0, column=1, padx=10, pady=5, sticky="nsew")
    ttk.Label(frm_weekday, text="Співробітники (будні):").grid(row=0, column=0, padx=5, pady=(5,2), sticky="w")
    default_weekday_var = tk.IntVar(value=0)
    chk_default_weekday = ttk.Checkbutton(frm_weekday, text="Default будні", variable=default_weekday_var,
                                           command=lambda: (text_weekday.delete("1.0", tk.END),
                                                            text_weekday.insert(tk.END, "\n".join(DEFAULT_WEEKDAY)),
                                                            update_weekday_start_dropdown()))
    chk_default_weekday.grid(row=1, column=0, padx=5, pady=2, sticky="w")
    text_weekday = tk.Text(frm_weekday, width=30, height=8)
    text_weekday.grid(row=2, column=0, padx=5, pady=5)
    ttk.Label(frm_weekday, text="Стартовий:").grid(row=3, column=0, padx=5, pady=(5,2), sticky="w")
    combobox_weekday_start = ttk.Combobox(frm_weekday, state="readonly", width=27)
    combobox_weekday_start.grid(row=4, column=0, padx=5, pady=2, sticky="w")
    update_weekday_start_dropdown()
    
    frame_workers.columnconfigure(0, weight=1)
    frame_workers.columnconfigure(1, weight=1)
    
    # Блок для відображення відпусток у головному вікні
    frame_vac_display = ttk.Frame(root, style="Custom.TFrame")
    frame_vac_display.grid(row=3, column=0, padx=5, pady=5, sticky="ew")
    ttk.Label(frame_vac_display, text="Відпустки:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
    text_vac_display = tk.Text(frame_vac_display, width=62, height=5, state="disabled")
    text_vac_display.grid(row=1, column=0, padx=5, pady=5)
    
    root.mainloop()
