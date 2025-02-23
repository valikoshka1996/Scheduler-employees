import calendar
import random
import re
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

def generate_color():
    """Генерує випадковий колір у форматі HEX."""
    return "{:06x}".format(random.randint(0, 0xFFFFFF)).upper()

def load_workers_from_file():
    """Завантажує працівників із вибраного файлу."""
    txt_files = [f for f in os.listdir() if f.endswith(".txt")]
    if not txt_files:
        print("Файли з працівниками не знайдено!")
        exit(1)
    
    print("Оберіть файл для вихідної зміни:")
    for i, file in enumerate(txt_files, 1):
        print(f"{i}. {file}")
    weekend_file = txt_files[int(input("Введіть номер: ")) - 1]
    
    print("Оберіть файл для буденної зміни:")
    for i, file in enumerate(txt_files, 1):
        print(f"{i}. {file}")
    weekday_file = txt_files[int(input("Введіть номер: ")) - 1]
    
    # Завантажуємо працівників у тій послідовності, в якій вони в файлі
    with open(weekend_file, encoding="utf-8") as f:
        weekend_workers = [line.strip() for line in f.readlines() if line.strip()]
    
    with open(weekday_file, encoding="utf-8") as f:
        weekday_workers = [line.strip() for line in f.readlines() if line.strip()]
    
    if not weekend_workers or not weekday_workers:
        print("Відсутні працівники у файлі!")
        exit(1)
    
    # Перевіряємо порядок елементів у кожному файлі
    print("Працівники для вихідної зміни:", weekend_workers)
    print("Працівники для буденної зміни:", weekday_workers)
    
    # Повертаємо два списки, в яких порядок елементів зберігається з файлів
    return weekend_workers, weekday_workers

WEEKDAYS = ["Понеділок", "Вівторок", "Середа", "Четвер", "П'ятниця", "Субота", "Неділя"]

def move_nth_to_first(lst, n):
    """Переміщує елемент за індексом n на перше місце у списку.

    Умови:
    - Якщо список містить 3 елементи - останній стає першим, зміщуючи інші.
    - Якщо список містить більше ніж 3 елементи - елемент за індексом n стає першим, а всі інші зміщуються далі.
    """
    if len(lst) == 3:
        # Якщо в списку 3 елементи - останній стає першим
        return [lst[-1]] + lst[:-1]
    elif len(lst) > 3:
        # Якщо в списку більше ніж 3 елементи - переміщуємо елемент за індексом n на перше місце
        if not (0 <= n < len(lst)):
            raise ValueError("Неправильний індекс")
        return [lst[n]] + lst[:n] + lst[n+1:]
    else:
        # Якщо елементів менше 3, нічого не змінюємо
        return lst


# Приклади використан
    
    return result
def generate_schedule(month, year, additional_weekends, weekend_workers, weekday_workers):
    """Генерує графік змін."""
    schedule = {}
    num_days = calendar.monthrange(year, month)[1]
    index_week = 3
    index_weeknd = 2
    workers_by_day_type = {"будній": weekday_workers.copy(), "вихідний": weekend_workers.copy()}
    weekday_queue = workers_by_day_type["будній"].copy()
    weekend_queue = workers_by_day_type["вихідний"].copy()
    
    for day in range(1, num_days + 1):
        current_date = datetime(year, month, day)
        is_weekend = current_date.weekday() >= 5 or day in additional_weekends
        day_type = "вихідний" if is_weekend else "будній"
        workers_queue = weekend_queue if is_weekend else weekday_queue
        
        if len(workers_queue) < 3:
            workers_queue = workers_by_day_type[day_type].copy()
        
        # Перший працівник на першій зміні, другий - на другій, третій - на третій
        schedule[day] = {
            "Перша зміна": workers_queue[0], 
            "Друга зміна": workers_queue[1],
            "Третя зміна": workers_queue[2] 
        }
        
        # Оновлення черги працівників (N + 1 замість N - 1)
        if is_weekend:
            index = index_weeknd
            weekend_queue = move_nth_to_first(workers_queue, index)
        else:
            index = index_week
            weekday_queue = move_nth_to_first(workers_queue, index)
            if index_week == len(weekday_workers) - 1:  
                index_week = 3
            else:
                index_week += 1
   
    return schedule

def save_to_excel(schedule, month, year):
    """Зберігає графік у файл Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    header = ["Дата"] + [f"{day:02}.{month:02}.{year}" for day in schedule.keys()]
    ws.append(header)
    
    ws.append(["День тижня"] + [WEEKDAYS[datetime(year, month, day).weekday()] for day in schedule.keys()])
    
    all_workers = {worker for day in schedule.values() for worker in day.values()}
    COLORS = {worker: generate_color() for worker in all_workers}
    
    for _ in range(4):
        ws.append(["Перша зміна"] + [schedule[day]["Перша зміна"] for day in schedule.keys()])
    for _ in range(4):
        ws.append(["Друга зміна"] + [schedule[day]["Друга зміна"] for day in schedule.keys()])
    for _ in range(4):
        ws.append(["Третя зміна"] + [schedule[day]["Третя зміна"] for day in schedule.keys()])
    
    for row in ws.iter_rows(min_row=3, min_col=2):
        for cell in row:
            if cell.value in COLORS:
                cell.fill = PatternFill(start_color=COLORS[cell.value], fill_type="solid")
    
    filename = f"Графік_{month:02}_{year}.xlsx"
    wb.save(filename)
    print(f"Графік збережено у {filename}")

def main():
    """Основна функція."""
    while True:
        month_year = input("Введіть місяць та рік у форматі MM.YYYY: ")
        if re.match(r"^(0[1-9]|1[0-2])\.\d{4}$", month_year):
            break
        print("Некоректний формат! Спробуйте ще раз.")
    
    month, year = map(int, month_year.split("."))
    additional_weekends = list(map(int, input("Введіть додаткові вихідні (через пробіл): ").split()))
    
    weekend_workers, weekday_workers = load_workers_from_file()
    schedule = generate_schedule(month, year, additional_weekends, weekend_workers, weekday_workers)
    save_to_excel(schedule, month, year)
    
if __name__ == "__main__":
    main()


input()